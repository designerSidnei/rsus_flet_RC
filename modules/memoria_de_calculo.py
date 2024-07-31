# Importar bibliotecas
import locale

import os
import shutil

import pandas as pd
import pywintypes
import xlwings as xw

from openpyxl.styles import PatternFill, Border, Side, Font, colors, Alignment

from modules.load_config import config_read

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

config = config_read()


# Função para verificar se existe a coluna
def find_column(df, column_names):
    for column_name in column_names:
        if column_name in df.columns:
            return column_name
    return None


# Função para salvar dataframe como Excel
def write_excel(df):
    with pd.ExcelWriter('planilha.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Carregar a planilha usando a biblioteca openpyxl
        workbook = writer.book
        worksheet = workbook['Sheet1']

        # Formatação do cabeçalho
        header_font = Font(color=colors.BLACK, bold=True, size=6)
        header_fill = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')
        header_border = Border(bottom=Side(border_style='thin'), left=Side(border_style='thin'),
                               right=Side(border_style='thin'), top=Side(border_style='thin'))
        header_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', shrinkToFit=True)

        # Formatação para os dados
        data_font = Font(color=colors.BLACK, size=6)
        data_fill = PatternFill(start_color=colors.WHITE, end_color=colors.WHITE, fill_type='solid')
        data_border = Border(bottom=Side(border_style='thin'), left=Side(border_style='thin'),
                             right=Side(border_style='thin'), top=Side(border_style='thin'))
        data_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center', shrinkToFit=True)

        # Aplicar estilo ao cabeçalho
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_alignment

        # Aplicar estilo aos dados
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_row=worksheet.max_row, max_col=worksheet.max_column):
            for cell in row:
                cell.font = data_font
                cell.fill = data_fill
                cell.border = data_border
                cell.alignment = data_alignment

        # Redimensionar a primeira coluna
        worksheet.column_dimensions['A'].width = 30

        # Mesclar ultima linha
        last_row = worksheet.max_row
        worksheet.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=6)


# Converter valores para reais brasileiros ou porcentagem
def convert_value(vlr_copart):
    if vlr_copart is not None:
        value_copart = '{:.0%}'.format(vlr_copart) if str(vlr_copart).startswith('0') else locale.currency(
            float(vlr_copart), grouping=True, symbol='R$')
    else:
        value_copart = '-'
    return value_copart


# Cálculo copart
def calcular_copart(vlr, vlr_procedimento, quantidade):
    operacao_copart = None

    if vlr is None:
        quantidade_x_copart_value = '-'
    elif str(vlr).startswith('0'):  # — > cálculo para valores em porcentagem
        operacao_copart = float(vlr_procedimento) * vlr
        quantidade_x_copart_value = locale.currency(float(operacao_copart), grouping=True, symbol='R$')
    else:  # — > cálculo para valores em reais
        operacao_copart = float(vlr) * float(quantidade)
        quantidade_x_copart_value = locale.currency(float(operacao_copart), grouping=True, symbol='R$')

    return quantidade_x_copart_value, operacao_copart

# Formata data para datetime
def formatar_data(competencia):
    if isinstance(competencia, float):
        competencia = str(int(competencia))

    if len(str(competencia)) == 5:
        data_formatada = f'01/0{str(competencia)[0]}/{str(competencia)[-4:]}'
        competencia_formatada = pd.to_datetime(data_formatada, format='%d/%m/%Y')
    elif len(str(competencia)) == 6:
        data_formatada = f'01/{str(competencia)[:2]}/{str(competencia)[-4:]}'
        competencia_formatada = pd.to_datetime(data_formatada, format='%d/%m/%Y')
    else:
        competencia_formatada = competencia
    return competencia_formatada

# Função para acrescentar os dados referente a copart
def process_copart(df, proced_col, quantidade_col, valor_total_col, tipo_proce_col, coparticipacao_col, vlr_copart_col,
                   quant_x_copart_col, copart, copart_sec, val_copart, val_copart_sec):
    lista_proced_consulta = config['nome_procedimento']['consulta']
    lista_proced_especial = config['nome_procedimento']['especial']
    lista_proced_ambulatorial = config['nome_procedimento']['ambulatorial']

    # Passa o valor da coparticipação para a coluna VALOR DA COPART
    def vlr_coparticipacao(valor):
        if valor != 'None' or valor != '':
            df.at[index, vlr_copart_col] = valor

    soma = 0
    # Itera sobre as células da coluna procedimento
    for index, row in df.iterrows():
        is_special = False
        # Passa os valores conforme a classificação dos procedimentos
        if any(word in row[proced_col] for word in lista_proced_especial):
            df.at[index, tipo_proce_col] = 'ESPECIAL'
            is_special = True
        elif any(word in row[proced_col] for word in lista_proced_ambulatorial):
            df.at[index, tipo_proce_col] = 'AMBULATORIAL'
            is_special = True
        elif any(word in row[proced_col] for word in lista_proced_consulta):
            df.at[index, tipo_proce_col] = 'CONSULTA'
            is_special = True
        else:
            df.at[index, tipo_proce_col] = 'BÁSICO'

        if is_special:
            vlr_coparticipacao(copart)
            quant_x_copart, opera_copart = calcular_copart(val_copart, row[valor_total_col], row[quantidade_col])
            df.at[index, quant_x_copart_col] = str(quant_x_copart)
            if opera_copart is not None:
                if opera_copart < row[valor_total_col]:
                    df.at[index, coparticipacao_col] = 'SIM'
                    soma += opera_copart
                else:
                    df.at[index, coparticipacao_col] = 'NÃO'
                    df.at[index, quant_x_copart_col] = '-'
            else:
                df.at[index, coparticipacao_col] = 'NÃO'
                df.at[index, quant_x_copart_col] = '-'
        else:
            vlr_coparticipacao(copart_sec)
            quant_x_copart, opera_copart = calcular_copart(val_copart_sec, row[valor_total_col], row[quantidade_col])
            df.at[index, quant_x_copart_col] = str(quant_x_copart)
            if opera_copart is not None:
                if opera_copart < row[valor_total_col]:
                    df.at[index, coparticipacao_col] = 'SIM'
                    soma += opera_copart
                else:
                    df.at[index, coparticipacao_col] = 'NÃO'
                    df.at[index, quant_x_copart_col] = '-'
            else:
                df.at[index, coparticipacao_col] = 'NÃO'
                df.at[index, quant_x_copart_col] = '-'
    return soma


async def process_rows(df, planilha_path):
    # Define os nomes das colunas
    procedimento_column = config['lista_procedimento']
    quantidade_column = config['lista_quantidade']
    valor_total_column = config['lista_total']
    memo_calc_column = ['MEMÓRIA DE CÁLCULO', 'MEMÓRIA DE CÁLCULO COPART']
    atendimento_column = config['lista_num_atend']
    competencia_column = config['lista_competencia']
    copart_column = config['lista_copart']
    copart_sec_column = config['lista_copart_sec']

    # Encontra as colunas
    procedimento = find_column(df, procedimento_column)
    quantidade = find_column(df, quantidade_column)
    valor_total = find_column(df, valor_total_column)
    memo_calc = find_column(df, memo_calc_column)
    numero_atendimento = find_column(df, atendimento_column)
    competencia = find_column(df, competencia_column)
    copart = find_column(df, copart_column)
    copart_sec = find_column(df, copart_sec_column)

    for index, row in df.iterrows():
        # Se as células dessa coluna tiver o valor 'EM ANEXO' pegar valores de número de atendimento e competência
        if str(row[memo_calc]).lower() == 'em anexo':
            num_atend = row[numero_atendimento]
            compet = row[competencia]

            # Indica quais linhas contém os valores de num_atend e compet
            filtro = (df[numero_atendimento] == num_atend) & (df[competencia] == compet)

            # Método loc[] do pandas para selecionar apenas as linhas que atendem ao filtro e as colunas desejadas
            df_filtrado = df.loc[filtro, [procedimento, quantidade, valor_total]]

            # Adicionar novas colunas com dados
            df_filtrado['TIPO DE PROCEDIMENTO'] = ''
            df_filtrado['COPARTICIPAÇÃO'] = ''
            df_filtrado['VALOR DA COPART'] = ''
            # df_filtrado['PROC X COPART'] = ''
            df_filtrado['QUANTIDADE PROC X COPART'] = ''

            # Formatar valores para porcentagem ou reais
            val_copart = row.get(copart) if not pd.isna(row.get(copart)) else None
            value_copart = convert_value(val_copart)

            val_copart_sec = row.get(copart_sec) if not pd.isna(row.get(copart_sec)) else None
            value_copart_sec = convert_value(val_copart_sec)

            soma = process_copart(df_filtrado, procedimento, quantidade, valor_total,
                                  'TIPO DE PROCEDIMENTO', 'COPARTICIPAÇÃO', 'VALOR DA COPART',
                                  'QUANTIDADE PROC X COPART', value_copart, value_copart_sec, val_copart,
                                  val_copart_sec)
            soma = locale.currency(float(soma), grouping=True, symbol='R$')

            # Formata a coluna do valor total (R$)
            df_filtrado[valor_total] = df_filtrado[valor_total].apply(
                lambda x: locale.currency(float(x), grouping=True, symbol='R$'))

            # Adiciona uma linha com o total
            df_filtrado.loc[-1] = ['TOTAL', '', '', '', '', '', soma]
            df_filtrado.index = df_filtrado.index + 1

            # Formata o dataframe para exportar
            df_filtrado = df_filtrado.fillna('')
            df_filtrado = df_filtrado.map(lambda x: str(x) if isinstance(x, (int, float)) else x)

            if planilha_path:
                dir_plan_path = os.path.dirname(planilha_path)
                pdf_path = f'{dir_plan_path}\\Memórias de cálculo'
                os.makedirs(pdf_path, exist_ok=True)
            else:
                return "Caminho da planilha não encontrado!"

            planilha_pdf = f'{int(num_atend)}.10 MEMÓRIA DE CÁLCULO C{formatar_data(compet).month}.pdf'
            planilha_memo = 'planilha.xlsx'

            # Cria uma nova aplicação Excel
            app = xw.App(visible=False)
            try:
                # Exporta a planilha formatada como planilha.xlsx
                write_excel(df_filtrado)
                xw_book = app.books.open(planilha_memo)
            except PermissionError:
                return ("Não pôde abrir a planilha.\nPlanilha não fechada da última vez devido a algum erro."
                        "\nFeche no gerenciador de tarefas e tente de novo.")
            except FileNotFoundError:
                return "A planilha.xlsx não foi encontrado"
            except Exception as e:
                return f"Error: {e}"

            xw_pdf = xw_book.sheets[0]
            try:
                xw_pdf.to_pdf(planilha_pdf)
            except pywintypes.com_error:
                xw_book.close()
                app.kill()
                return (f"O documento não foi salvo. Talvez esteja aberto ou pode ter ocorrido um erro "
                        "durante a gravação.")
            except Exception as e:
                return f"Erro ao salvar pdf: {e}"

            xw_book.close()
            app.kill()

            try:
                shutil.move(planilha_pdf, pdf_path)
            except FileExistsError:
                return "Arquivo já existe na pasta"
            except Exception as e:
                return f"Error: {e}"

    return "Concluído com sucesso"
