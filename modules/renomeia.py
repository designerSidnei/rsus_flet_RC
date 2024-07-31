import datetime
from pathlib import Path
from openpyxl import load_workbook
from modules.procura_num_coluna import find_col

from modules.load_config import config_read


config = config_read()
competence_columns = config['comp']
attendance_number_columns = config['num_atend']


async def caminho_arquivos(plan_path, arquivos):
    plan = load_workbook(plan_path).active
    #arquivos = arquivos.split(',')
    competence_col = find_col(plan, competence_columns)
    attendance_number_col = find_col(plan, attendance_number_columns)
    rows = [(plan.cell(row=i, column=competence_col).value, plan.cell(row=i, column=attendance_number_col).value) for i
            in range(2, plan.max_row + 1)]

    new_name_list = []
    for competence, attendance_number in rows:
        if isinstance(competence, datetime.date):
            competence = competence.strftime('%m')
        else:
            competence = str(competence)
            competence = competence[:-4]

        if competence.startswith('0'):
            competence = competence[1:]

        attendance_number = str(attendance_number)[:14]
        new_name = f'{attendance_number}.0 C{competence}.pdf'

        for file in arquivos:
            if file:
                file_path = Path(file.strip())
                if not file_path.name.startswith(attendance_number):
                    new_file_path = file_path.parent / new_name
                    file_path.rename(new_file_path)
                    new_name_list.append(str(new_file_path.name))
                    arquivos.remove(file)
                    break

    return new_name_list
